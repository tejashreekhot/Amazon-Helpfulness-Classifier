#__author__ = 'sanjanaagarwal'
from openpyxl import load_workbook
wb = load_workbook(filename = 'HyperboleHalf(done).xlsx', read_only= True, use_iterators=True) #Name of the workbook. read_only to be used,
# if only reading
sheetName = wb['HH']
#ws1 = wb.create_sheet(title="HHTry" )
#cell_range = sheetName['A:B']
fulllist=[]
print sheetName.iter_rows()
for row in sheetName.rows:
    r=[]
    for cell in row:
        r.append(cell.value)
    fulllist.append(r)
listwewant=[(x[1], x[0].lower()) for x in fulllist]
print listwewant

    #print("\n")
   # print sheetName.columns[0].value
    #print sheetName.columns[10].value

